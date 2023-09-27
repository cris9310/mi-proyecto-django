from typing import Any, Dict
from django.shortcuts import render
import datetime
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook
from decimal import *
import warnings


from django.shortcuts import render, HttpResponseRedirect, HttpResponse
from django.contrib import messages
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.db.models.functions import Coalesce
from django.db.models import F 

from .funciones import  generador
# Create your views here.

from .models import *
from Applications.agenda.models import *
from Applications.financiero.models import *
from django.views.generic import (TemplateView, 
FormView, CreateView, DeleteView, UpdateView,
DetailView, ListView, View)


from .forms import *
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Count
from .choices import EST_STUDENT
import pandas as pd
import re


#Esta vista se tiene que modificar, teniendo en cuenta el perfil de quien se está logueando
class Dashboard(ListView):
    model= Estudiante
    template_name = 'Academico/Dashboard/dashboard.html'
    context_object_name = 'dashboard'

    def get_context_data(self, **kwargs):
        context = super(Dashboard, self).get_context_data(**kwargs)
        context['student'] = Estudiante.objects.filter(is_active=True).count()
        context['docentes'] = Docente.objects.filter(is_active=True).count()
        user_id = []
        all_events_user = Evento.objects.filter(user_id=self.request.user)
        
        for i in all_events_user:
            rr = {'id':i.id, 'title': i.title, 'descripcion': i.descripcion, 'fecha_inicial': i.start_time, 'fecha_final': i.end_time, 'color':i.color}
            user_id.append(rr)
        context["eventos"] = user_id

        return context
        
    def get_queryset(self):
        estudiante_calc = []
        if Estudiante.objects.filter(is_active=True).count() != 0:
           mujeres = round((Estudiante.objects.filter(is_active=True, sexo="Femenino").count()/Estudiante.objects.filter(is_active=True).count())*100,0)
           hombres= round((Estudiante.objects.filter(is_active=True, sexo="Masculino").count()/Estudiante.objects.filter(is_active=True).count())*100,0)
           otros= round((Estudiante.objects.filter(is_active=True, sexo="Otros").count()/Estudiante.objects.filter(is_active=True).count())*100,0)
           calculo = {'mujeres':str(mujeres), 'hombres':str(hombres), 'otros':str(otros)}
           estudiante_calc.append(calculo)
        else:
           mujeres = round((Estudiante.objects.filter(is_active=True, sexo="Femenino").count()/1)*100,0)
           hombres= round((Estudiante.objects.filter(is_active=True, sexo="Masculino").count()/1)*100,0)
           otros= round((Estudiante.objects.filter(is_active=True, sexo="Otros").count()/1)*100,0)
           calculo = {'mujeres':str(mujeres), 'hombres':str(hombres), 'otros':str(otros)}
           estudiante_calc.append(calculo)

        return estudiante_calc



class UserCreateView(FormView):
    form_class = UserRegisterForm
    template_name = 'Academico/usuarios/register_user.html'
    success_url =  reverse_lazy('academico_app:create-user')

    def form_valid(self, form):
        crea_user=User.objects.create_user(
            tipe = form.cleaned_data['tipe'],
            codigo = form.cleaned_data['codigo'],
            username= form.cleaned_data['username'],
            email= form.cleaned_data['email'],
            password= form.cleaned_data['password1'],
            nombres= form.cleaned_data['nombres'],
            apellidos= form.cleaned_data['apellidos'],
            is_superuser=False,
            is_active=True,
            is_staff= False,

        )
            
        return HttpResponseRedirect(self.success_url)
    

## Vista ok
class ConfigTemplateView(TemplateView):
    template_name = "Academico/settings/setings.html"

class PeriodoCreateView(FormView):
    form_class = PeriodoForm
    template_name = 'Academico/settings/Periodos/register_periodo.html'
    success_url =  reverse_lazy('academico_app:config')

    def form_valid(self, form):
        periodo=str(form.cleaned_data['periodo']),
        create_peri=Periodos.objects.create_periodo(
            periodo[0],
           
        )
        return HttpResponseRedirect(self.success_url)

class PeriodoListView(ListView):
    model= Periodos
    template_name = 'Academico/settings/Periodos/list_periodos.html'
    context_object_name = 'periodos'
    paginate_by = 5
    
    def get_queryset(self):

        data_periodos = Periodos.objects.all()
        data=[]
        for i in data_periodos:
            if Materias.objects.filter(periodo_id=i.id):
                data_json = {"id":i.id,"periodo":i.periodo,"estado":True}
                data.append(data_json)
            else:
                data_json = {"id":i.id,"periodo":i.periodo,"estado":False}
                data.append(data_json)

        return data

class PeriodoDeleteView(DeleteView):
    template_name = 'Academico/settings/Periodos/detele_periodo.html'
    model = Periodos
    success_url = reverse_lazy('academico_app:list-periodo')
    

class ProgramaCreateView(FormView):
    form_class = ProgramaForm
    template_name = 'Academico/settings/Programas/register_program.html'
    success_url =  reverse_lazy('academico_app:config')

    
    def form_valid(self, form):

        crear_programa = Programas.objects.crear(
            tipe = form.cleaned_data['tipe'],
            cod_prog= Programas.objects.code_programas(),
            matricula = form.cleaned_data['matricula'],
            cuota_valor= form.cleaned_data['cuota_valor'],
            cuotas= form.cleaned_data['cuotas'],
            costo= form.cleaned_data['costo'],
            programa_name= form.cleaned_data['programa_name'],
            aceptado = form.cleaned_data['aceptado']
            
        )

        return HttpResponseRedirect(self.success_url)


class ProgramaDetailView(DetailView):
    template_name = 'Academico/settings/Programas/detail_program.html'
    model = Programas

class ProgramaDeleteView(DeleteView):
    template_name = 'Academico/settings/Programas/delete_program.html'
    model = Programas
    success_url = reverse_lazy('academico_app:list-program')

class ProgramaUpdateView(UpdateView):
    template_name = 'Academico/settings/Programas/update_program.html'
    model = Programas
    form_class =ProgramaForm
    success_url = reverse_lazy('academico_app:list-program')

class Programalistview(ListView):
    model = Programas
    template_name = 'Academico/settings/Programas/list_programas.html'
    context_object_name = 'programas'
    paginate_by = 10

    def get_queryset(self):
        kword = self.request.GET.get('kword', '')
        data_program = []
        if kword:
            data_prin= Programas.objects.filtrar_buscador(kword)
            for i in data_prin:
                if Materias.objects.filter(programa_id=i.id).exists():
                    data_json = {"pk":i.id,"codigo":i.cod_prog,"programa":i.programa_name,"estado":True, "is_active":i.is_active}
                    data_program.append(data_json)
                else:
                    data_json = {"pk":i.id,"codigo":i.cod_prog,"programa":i.programa_name,"estado":False, "is_active":i.is_active}
                    data_program.append(data_json)
            queryset =data_program

        
        else:

            data_prin= Programas.objects.all()
            for i in data_prin:
                if Materias.objects.filter(programa_id=i.id).exists():
                    data_json = {"pk":i.id,"codigo":i.cod_prog,"programa":i.programa_name,"estado":True, "is_active":i.is_active}
                    data_program.append(data_json)
                else:
                    data_json = {"pk":i.id,"codigo":i.cod_prog,"programa":i.programa_name,"estado":False, "is_active":i.is_active}
                    data_program.append(data_json)
            queryset =data_program

        return queryset

class CrearPensum(FormView):
    form_class = PensumRegisterForm
    template_name = 'Academico/settings/Programas/register_pensum.html'
    success_url =  reverse_lazy('academico_app:create-pensum')

    def form_valid(self, form):
        malla = form.cleaned_data['malla']
        programa = form.cleaned_data['programa']

        if Pensum.objects.filter(malla = malla, programa = programa):
            messages.warning(self.request,'El pensum que intenta crear ya existe para este programa, por favor Verifique y vuelva a intentarlo.')
            return HttpResponseRedirect(self.success_url)

        else:
            crear_pensum = Pensum.objects.create(
                malla=malla,
                programa=programa
            )
            return HttpResponseRedirect(self.success_url)


#Vista ok, se encarga de listar docentes, filtra activos, inactivos y todos.
class Teacherlistview(ListView):
    model = Docente
    template_name = 'Academico/Docentes/list_teacher.html'
    context_object_name = 'teacher'

    def get_context_data(self, **kwargs):
        context = super(Teacherlistview, self).get_context_data(**kwargs)
        context['todos'] = Docente.objects.all().count()
        context['activos'] = Docente.objects.filter(is_active=True).count()
        context['inactivos'] = Docente.objects.filter(is_active=False).count()
        return context

    def get_queryset(self):
        kword = self.request.GET.get('kword', '')
        order = self.request.GET.get('order', '')
        data_teacher = []
        if kword or order:
            data_prin = Docente.objects.filtrar_teacher(kword, order)
            for i in data_prin:

                if Materias.objects.filter(docente_id=i.id).exists():
                    data_json = {"pk":i.id,"codigo":i.codigo,"nombres":i.nombres,"apellidos":i.apellidos,"estado":True, "is_active":i.is_active}
                    data_teacher.append(data_json)
                else:
                    data_json = {"pk":i.id,"codigo":i.codigo,"nombres":i.nombres,"apellidos":i.apellidos,"estado":False,"is_active":i.is_active}
                    data_teacher.append(data_json)
            queryset =data_teacher
            return queryset

        
        else:
            data_prin=Docente.objects.all()
            for i in data_prin:
                if Materias.objects.filter(docente_id=i.id).exists():
                    data_json = {"pk":i.id,"codigo":i.codigo,"nombres":i.nombres,"apellidos":i.apellidos,"estado":True,"is_active":i.is_active}
                    data_teacher.append(data_json)
                else:
                    data_json = {"pk":i.id,"codigo":i.codigo,"nombres":i.nombres,"apellidos":i.apellidos,"estado":False,"is_active":i.is_active}
                    data_teacher.append(data_json)
            queryset =data_teacher
            return queryset


#Vista que sirve crear docentes, se encuentra ok
class TeacherCreateView(CreateView):
    model = Docente
    form_class = TeacherForm
    template_name = 'Academico/Docentes/register_teacher.html'
    success_url =  reverse_lazy('academico_app:list-teacher')

    def post(self, request, *args, **kwargs):

        formulario = TeacherForm(self.request.POST)
        if formulario.is_valid():
            formulario.save()
            crea_user=User.objects.create_user(
                tipe = CatalogsTypesRol.objects.get(rol="Docente"),
                codigo = formulario.cleaned_data.get('codigo'),
                username= formulario.cleaned_data.get('username'),
                email= formulario.cleaned_data.get('email'),
                password=Docente.objects.get_secret("RANDOM"),
                nombres= formulario.cleaned_data.get('nombres'),
                apellidos= formulario.cleaned_data.get('apellidos'),
                is_superuser=False,
                is_active=True,
                is_staff= False,
            )
            mensaje = f'{self.model.__name__} registrado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar'
            error = formulario.errors
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 400
            return response

#Vista para actualizar la información y el usuario del docente. Se encuentra ok.
class TeacherUpdateView(UpdateView):
    model = Docente
    template_name = 'Academico/Docentes/update_teacher.html'
    form_class = TeacherForm
    success_url = reverse_lazy('academico_app:list-teacher')

    def post(self, request, *args, **kwargs):
        self.object=self.get_object
        cod_teacher=Docente.objects.get(pk=self.kwargs['pk']).codigo
        teacher_c=self.model.objects.get(codigo=cod_teacher)
        form = TeacherForm(request.POST, instance=teacher_c)

        if form.is_valid():
            form.save()
            crea_user=User.objects.filter(
                codigo =cod_teacher
            ).update(
                username= form.cleaned_data['username'],
                email= form.cleaned_data['email'],
                nombres= form.cleaned_data['nombres'],
                apellidos= form.cleaned_data['apellidos'],
                updated_at = datetime.now()
            )
            mensaje = f'{self.model.__name__} actualizado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar'
            error = form.errors
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 400
            return response
            

# Vista ok, muestra el detalle del docente ok
class TeacherDetailView(DetailView):
    template_name = 'Academico/Docentes/detail_teacher.html'
    model = Docente



#Esta vista se enceuntra ok, muestra el detalle de las asignaturas que tiene el docente. ok
class TeacherTopicsListview(ListView):
    model = Docente
    second_model = Materias
    template_name = 'Academico/Docentes/list_materias_teacher.html'
    context_object_name = 'teacherTp'

    def get_context_data(self, **kwargs):
        cod = self.kwargs['pk']
        data = []
        context = super(TeacherTopicsListview, self).get_context_data(**kwargs)
        context['topics'] = Docente.objects.filter(pk=cod)
        context['Tmaterias'] = Materias.objects.filter(docente=cod).count()
        return context
    

    def get_queryset(self):

        cod = self.kwargs['pk']
        kword = self.request.GET.get('kword', '')
        data = []
        if kword:
           for i in Materias.objects.filtrar_buscador_Teacher_topic(kword, cod):
                try:
                    info = Banner.objects.values("materia__codigo").annotate(total=Coalesce(Count("materia"),0)).filter(materia__codigo=i.codigo)
                    info1 = {"pk": i.pk, "codigo": i.codigo, "programa": i.programa, "materia": i.materia, "total":info[0]['total'], "periodo": i.periodo }
                    data.append(info1)

                except:
                    info1 = {"pk": i.pk, "codigo": i.codigo, "programa": i.programa, "materia": i.materia, "total": 0, "periodo": i.periodo}
                    data.append(info1)
        else:
            for i in Materias.objects.filter(docente=cod):
                try:
                    info = Banner.objects.values("materia__codigo").annotate(total=Coalesce(Count("materia"),0)).filter(materia__codigo=i.codigo)
                    info1 = {"pk": i.pk, "codigo": i.codigo, "programa": i.programa, "materia": i.materia, "total":info[0]['total'], "periodo": i.periodo }
                    data.append(info1)

                except:
                    info1 = {"pk": i.pk, "codigo": i.codigo, "programa": i.programa, "materia": i.materia, "total": 0, "periodo": i.periodo}
                    data.append(info1)
        return data

# Vista que muestra el detalle de las notas del salon, se encuentra ok

class TeacherNotesListview(ListView):
    model = Docente
    second_model = Materias
    template_name = 'Academico/Docentes/list_notas_teacher.html'
    context_object_name = 'teacherTp'

    def get_context_data(self, **kwargs):
        cod = self.kwargs['pk']
        data = Habilitaciones.objects.latest("id").id
        context = super(TeacherNotesListview, self).get_context_data(**kwargs)
        context['topics'] = Materias.objects.filter(pk=cod)
        context['total'] = Banner.objects.filter(materia=cod).count()
        context['corte'] = Habilitaciones.objects.filter(id=data)
        return context
    
    def get_queryset(self):

        cod = self.kwargs['pk']
        kword = self.request.GET.get('kword', '')
        if kword:
           queryset = Banner.objects.filtrar_buscador_Teacher_notes(kword, cod) 
        else:
            queryset = Banner.objects.filter(materia=cod)
            
        return queryset.order_by('cod_student')

# Vista que habilita o inhabilita docentes, se enceuntra ok
class TeacherHabilitView(UpdateView):
    model = Docente
    template_name = 'Academico/Docentes/update_habilitar.html'
    fields = '__all__'
    success_url = reverse_lazy('academico_app:list-teacher')

    def post(self, request, *args, **kwargs):
        accion = self.request.POST.get('accion')
        if accion == "inhabilitar":
            object1= User.objects.filter(
                pk=Docente.objects.get(pk=self.kwargs['pk']).codigo
                ).update(is_active=False)

            object2= Docente.objects.filter(
                pk=self.kwargs['pk']
                ).update(is_active=False)
            
            messages.success(self.request,'El docente ha sido inhabilitado correctamente')
            return HttpResponseRedirect(self.success_url)
        
        else:
            object1= User.objects.filter(
                pk=Docente.objects.get(pk=self.kwargs['pk']).codigo
                ).update(is_active=True)

            object2= Docente.objects.filter(
                pk=self.kwargs['pk']
                ).update(is_active=True)
            
            messages.success(self.request,'El docente ha sido habilitado correctamente')
            return HttpResponseRedirect(self.success_url)

class TeacherDeleteView(DeleteView):
    template_name = 'Academico/Docentes/delete_teacher.html'
    model = Docente
    success_url = reverse_lazy('academico_app:list-teacher')

    def post(self, request, *args, **kwargs):
        
        data_delete = self.request.POST
        teacher_delete = Docente.objects.filter(codigo = data_delete["codigo"]).delete()
        user_delete = User.objects.filter(codigo = data_delete["codigo"]).delete()
            
        messages.success(self.request,'El docente ha sido eliminado correctamente')
        return HttpResponseRedirect( reverse_lazy('academico_app:list-teacher'))


class MateriasCreateView(FormView):
    form_class = MateriasForm
    template_name = 'Academico/settings/Materias/register_materia.html'
    success_url =  reverse_lazy('academico_app:list-topic')

    def form_valid(self, form):
        
        form = MateriasForm(self.request.POST)
        def code_materias_generator():

            try:
                filtro = Materias.objects.latest('id').codigo
                cod_asign = int(filtro) + 1
                return cod_asign
            except:
                año = datetime.now()
                cod_asign= str(año.year) + str("01")
                return str(cod_asign)

        codigo = code_materias_generator()
            
        if form.is_valid():
            crea_materias = Materias.objects.create(
                codigo = codigo,
                materia = form.cleaned_data['materia'],
                programa = form.cleaned_data['programa'],
                docente = form.cleaned_data['docente'],
                periodo = form.cleaned_data['periodo'],
                sede = form.cleaned_data['sede'],
                pensum_asig = form.cleaned_data['pensum_asig']

            )
            
        
        return HttpResponseRedirect(self.success_url)

class MateriaUpdateView(UpdateView):
    template_name = 'Academico/settings/Materias/update_topic.html'
    model = Materias
    form_class = MateriasForm
    success_url = reverse_lazy('academico_app:list-topic')


#Mustra el detalle de las materias en un modal se encuentra ok
class MateriaDetailView(DetailView):
    template_name = 'Academico/settings/Materias/detail_materia.html'
    model = Materias

class Materialistview(ListView):
    model = Materias
    template_name = 'Academico/settings/Materias/list_topic.html'
    context_object_name = 'materias'
    paginate_by = 10

    def get_queryset(self):
        kword = self.request.GET.get('kword', '')

        if kword:
            queryset = Materias.objects.filtrar_buscador_materias(kword)
        
        else:
            queryset=Materias.objects.all()

        return queryset

class MateriaDeleteView(DeleteView):
    template_name = 'Academico/settings/Materias/delete_topic.html'
    model = Materias
    success_url = reverse_lazy('academico_app:list-topic')

class Getpunsumview(View):
    
    def get(self, request, *args, **kwargs):

        carrera = self.request.GET.get('estado_id')
        materia = self.request.GET.get('materia')
        options = '<option value="" selected="selected">---------</option>'
        if carrera:
            carrera_todo = Pensum.objects.filter(programa = carrera)
            for i in carrera_todo:
                options += '<option value="%s">%s</option>'% (
                i.pk,
                i.malla
            )
            response = {}
            response['malla'] = options
            return JsonResponse(response)

        elif materia: 
            carrera_todo = Inventario.objects.filter(pensum_asig_inv = materia)
            for i in carrera_todo:
                options += '<option value="%s">%s</option>'% (
                i.pk,
                i.nombre_materia
            )
            response = {}
            response['materia'] = options
            return JsonResponse(response)


        else:
            response = {}
            return JsonResponse(response)

class InventarioCreateView(FormView):
    form_class = InventarioRegisterForm
    template_name = 'Academico/settings/Materias/register_inventario.html'
    success_url =  reverse_lazy('academico_app:list-topic')

    def form_valid(self, form):

        create_inventario= Inventario.objects.create(
            nombre_materia = form.cleaned_data['nombre_materia'],
            programa = form.cleaned_data['programa'],
            pensum_asig_inv = form.cleaned_data['pensum_asig_inv'],


        )
        return HttpResponseRedirect(self.success_url)



class MateriaDetailView(DetailView):
    template_name = 'Academico/settings/Materias/detail_materia.html'
    model = Materias


##Con esta vista se muestran los datos del usuario logueado okkkkkkkk
class ProfileDetailView(ListView):
    model = User
    template_name = 'Academico/usuarios/profile.html'
    context_object_name = 'consulta'

    def get_queryset(self):


        if self.request.user.tipe_id == 4 or self.request.user.tipe_id == 5 :
            if self.request.user.tipe_id == 5:
               queryset= Docente.objects.filter(codigo=self.request.user.pk)
               return queryset
            else:
                queryset= Estudiante.objects.filter(codigo=self.request.user.pk)
                return queryset
        else:
            queryset= User.objects.filter(codigo=self.request.user.pk)
            return queryset
    

### CRUD de estudiantes


#Vista que sirve para la creación de estudiantes de manera individual, se encuentra ok
class StudentCreateView(CreateView):
    model = Estudiante
    form_class = StudentRegisterForm
    template_name = 'Academico/Estudiantes/register_student.html'
    success_url =  reverse_lazy('academico_app:list-student')

    def post(self, request, *args, **kwargs):
  
        form = StudentRegisterForm(self.request.POST)
        if form.is_valid():
            inv_list=[]
            #Generador de código para facturas
            codigo_invoice= Facturas.objects.code_invoice()
            #Generador de código para un nuevo estudiante
            codigo = Estudiante.objects.code_generatorr()
            #Creando estudiantes en la bd de estudiante
            create_student= Estudiante.objects.create(
                codigo=codigo,
                tDocument=CatalogsTypesDocuement.objects.get(nombre=form.cleaned_data.get("tDocument")),
                apellidos=form.cleaned_data.get('apellidos'),
                username= form.cleaned_data.get('username'),
                direccion=form.cleaned_data.get('direccion'),
                nacimiento=form.cleaned_data.get('nacimiento'),
                sexo=form.cleaned_data.get('sexo'),
                telefono=form.cleaned_data.get('telefono'),
                email=form.cleaned_data.get('email'),
                cedula=form.cleaned_data.get('cedula'),
                nombre=form.cleaned_data.get('nombre'),
                nacionalidad=form.cleaned_data.get('nacionalidad'),
                carrera=Programas.objects.get(programa_name=form.cleaned_data.get('carrera')),
                sede =CatalogsSede.objects.get(sede=form.cleaned_data.get('sede')),
                costo_cierre=Programas.objects.get(programa_name=form.cleaned_data.get('carrera')).costo,
                nombre_acudiente=form.cleaned_data.get('nombre_acudiente'),
                apellidos_acudiente=form.cleaned_data.get('apellidos_acudiente'),
                telefono_acudiente=form.cleaned_data.get('telefono_acudiente'),
                cedula_acudiente=form.cleaned_data.get('cedula_acudiente'),
                periodo_matriculado = Periodos.objects.get(periodo=form.cleaned_data.get('periodo_matriculado')),
                pensum_asig =Pensum.objects.get(id=form['pensum_asig'].value()),
            )
            #Creamos usuario en la BD
            crea_user=User.objects.create_user(
                tipe = CatalogsTypesRol.objects.get(rol="Estudiante"),
                codigo = codigo,
                username= form.cleaned_data.get('username'),
                email= form.cleaned_data.get('email'),
                password= Estudiante.objects.get_secret("RANDOM"),
                nombres= form.cleaned_data.get('nombre'),
                apellidos= form.cleaned_data.get('apellidos'),
                is_superuser=False,
                is_active=True,
                is_staff= False,
            )
            #Generamos las facturas relacionadas con el estudiante, hace referencia a matrícula y a las facturas de las pensiones mensuales
            l= ["Matricula", "Derechos de grado"]
            for i in l:
                matricula= Facturas.objects.create(
                        user=User.objects.get(username=form.cleaned_data.get('username')),
                        codigo=codigo_invoice,
                        email=form.cleaned_data.get('email'),
                        monto =Programas.objects.get(programa_name=form.cleaned_data.get('carrera')).matricula,
                        descripcion=i,
                        estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                )
            for i in range(int(Programas.objects.get(programa_name=form.cleaned_data.get('carrera')).cuotas)):
                inv_list.append(Facturas(
                    user=User.objects.get(username=form.cleaned_data.get('username')),
                    codigo=int(Facturas.objects.code_invoice()) + i,
                    email=form.cleaned_data.get('email'),
                    monto =Programas.objects.get(programa_name=form.cleaned_data.get('carrera')).cuota_valor,
                    descripcion="Mensualidad número "+ str(i+1),
                    estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),

                ))
            Facturas.objects.bulk_create(inv_list)

            mensaje = f'{self.model.__name__} registrado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar'
            error = form.errors
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 400
            return response

class Vernotas(ListView):
    model = Banner
    second_model = Estudiante
    template_name = 'Academico/Estudiantes/list_notas.html' 
    context_object_name = 'notas'

    
    def get_context_data(self, **kwargs):
        cod = self.kwargs['pk']
        context = super(Vernotas, self).get_context_data(**kwargs)
        context['student'] = Estudiante.objects.filter(pk=cod)
        context['progreso'] = int(int(Banner.objects.filter(cod_student=cod, promedio__gt=2.99).count())/int(Programas.objects.get(
            id = Estudiante.objects.get(pk=cod).carrera_id
        ).aceptado)*100)
        return context
    

    def get_queryset(self):

        cod = self.kwargs['pk']
        queryset = Banner.objects.filter(cod_student=cod)
        return queryset

class AsignamateriaPreviewView(ListView):
    model = Estudiante
    template_name = "Academico/Estudiantes/enrolled_student.html"
    context_object_name = 'programas'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        info = Estudiante.objects.get(id=self.kwargs['pk']).carrera
        context['estudiante']= Estudiante.objects.get(id=self.kwargs['pk'])
        context['programa']= Programas.objects.get(id=info.id)
        context['pensum'] = Estudiante.objects.get(id=self.kwargs['pk']).pensum_asig
        context['periodo'] = Estudiante.objects.get(id=self.kwargs['pk']).periodo_matriculado
        return context
    
    def get_queryset(self, *args, **kwargs):    
        id_estudiante = self.kwargs['pk']
        alumno = Estudiante.objects.get(id=id_estudiante).carrera
        pensum_asig = Estudiante.objects.get(id=id_estudiante).pensum_asig
        periodo_asig = Estudiante.objects.get(id=id_estudiante).periodo_matriculado
        sede_asig = Estudiante.objects.get(id=id_estudiante).sede
        queryset1 = Materias.objects.filter(programa_id = alumno.id, pensum_asig_id= pensum_asig.id, periodo_id = periodo_asig.id, materia_id= sede_asig.id )
        
        return queryset1


#Esta vista lista a todos los estudiantes que se han creado, se enceuntra ok.
class Studentlistview(ListView):
    model = Estudiante
    template_name = 'Academico/Estudiantes/list_student.html'
    context_object_name = 'student'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['programas'] = Programas.objects.all()
        context['masivo'] = Estudiante.objects.filter(masivo=True).count()
        return context

    def get_queryset(self):
        kword = self.request.GET.get('kword', '')
        order = self.request.GET.get('order', '')
        if kword or order:
            queryset = Estudiante.objects.filtrar_student(kword, order).exclude(masivo=True)
        else:
            queryset=Estudiante.objects.filter(is_active=True).exclude(masivo=True)

        return queryset

class StudentCargueListview(ListView):
    model = Estudiante
    template_name = 'Academico/Estudiantes/masivos.html'
    context_object_name = 'student'
    paginate_by = 15

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['programas'] = Programas.objects.all()
        return context

    def get_queryset(self):
        kword = self.request.GET.get('kword', '')
        order = self.request.GET.get('order', '')
        if kword or order:
            queryset = Estudiante.objects.filtrar_student(kword, order).exclude(masivo=False)
        else:
            queryset=Estudiante.objects.filter(is_active=True).exclude(masivo=False)

        return queryset

class StudentDetailView(DetailView):
    template_name = 'Academico/Estudiantes/detail_student.html'
    model = Estudiante

class StudentUpdateView(UpdateView):
    model = Estudiante
    template_name = 'Academico/Estudiantes/update_student.html'
    form_class = StudentRegisterForm
    success_url =reverse_lazy('academico_app:list-student') 
    

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        cod_estudiante=Estudiante.objects.get(pk=self.kwargs['pk']).codigo
        estudiante_c=self.model.objects.get(codigo=cod_estudiante)
        form = StudentRegisterForm(request.POST, instance=estudiante_c)
        

        if form.is_valid():
            form.save()
            crea_user=User.objects.filter(
                codigo =cod_estudiante
            ).update(username=form.cleaned_data['username'],
                email = form.cleaned_data['email'])
            act = Estudiante.objects.filter(pk=self.kwargs['pk']).update(masivo = False)
        return HttpResponseRedirect(self.success_url)

#Vista para cargar estudiantes de forma masiva se encuentra ok
class StudentAsigView(View):

    def get(self, request, *args, **kwargs):
        form = StudentAsigMate()
        context = {"form" : form} 
        return render(request, r"Academico\Estudiantes\asig_topics.html", context)

    def post(self, request, *args, **kwargs):

        # Declarando variables
        conteo = 0
        fecha_actual = date.today()
        
        # Declarando listas
        mensaje1 = []
        data_list = []
        data_fac_matricula= []
        l= ["Matricula", "Derechos de grado"]

        #Leyendo el archivo que recibimos desde el form
        new_studentData = self.request.FILES['carga']
        
         # zona de verificación de datos en el archivo

        if new_studentData.name == 'cargue_estudiantes.xlsx' :
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            new_student = pd.read_excel(new_studentData, sheet_name="Plantilla", engine='openpyxl' )
            new_student['NACIMIENTO'] = pd.to_datetime(new_student['NACIMIENTO'])
            new_student['DUPLICADO'] = new_student.duplicated()

            #Validamos que no existan valores nulos y en blancos en el archivo.
            miss_values_count = new_student.isnull().sum()
            miss_values_count = miss_values_count[miss_values_count != 0]

            #Generamos código para el estudiante   
            codigo = Estudiante.objects.code_generatorr()

            #comparando los encabezados para ver si está correcta la estructura
            for i, a in zip(list(new_student), EST_STUDENT):
                if i != a:
                    break
        
            if miss_values_count.shape[0] or i != a :
                if  miss_values_count.shape[0]:
                    if miss_values_count.shape[0] == 1:
                        mensaje1.append({"error":'El archivo tiene '+ str(miss_values_count.shape[0]) + ' columna sin datos, por favor, revise'})
                        response = JsonResponse(mensaje1, safe= False)
                        response.status_code = 400
                        return response 
                    else:
                        mensaje1.append({"error":'El archivo tiene '+ str(miss_values_count.shape[0]) + ' columnas sin datos, por favor, revise'})
                        response = JsonResponse(mensaje1, safe= False)
                        response.status_code = 400
                        return response 

                else:
                    mensaje1.append({"error":'La columna con encabezado '+ i + ' en su archivo, no es válido, verifique el archivo y vuelva a cargarlo'})
                    response = JsonResponse(mensaje1, safe= False)
                    response.status_code = 400
                    return response 
            
            elif len(new_student) > 35 or len(new_student) == 0:
                if len(new_student) > 35:
                    mensaje1.append({"error":'Recuerde que no puede cargar más de 35 estudiantes a la vez, en este archivo encontramos: ' + str(len(new_student))})
                    response = JsonResponse(mensaje1, safe= False)
                    response.status_code = 400
                    return response 
                else:
                    mensaje1.append({"error":'No encontramos estudiantes para cargar'})
                    response = JsonResponse(mensaje1, safe= False)
                    response.status_code = 400
                    return response
            
            else:
                #Validamos los campos
                
                for i in range(len(new_student)):
                    if  new_student['DUPLICADO'][i] == True:
                        conteo = 1 + conteo
                        mensaje1.append({"error": "Encontramos filas duplicadas para el usuario: " + new_student['USERNAME'][i] + " verifique la información"})
                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    if User.objects.filter(username=  new_student['USERNAME'][i]):
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El nombre de usuario " + new_student['USERNAME'][i] + " ya existe, por favor cambie este username y vuelva a intentarlo."})
                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):

                    res = " " in new_student['USERNAME'][i]
                    if res:
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El usuario " + new_student['USERNAME'][i] + " contiene espacios en blanco, por favor verifique."})
                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    if Programas.objects.filter(programa_name= new_student['CARRERA_ID'][i]):
                        conteo = 0 + conteo
                    else:
                        conteo = 1 + conteo
                        mensaje1.append({"error": "La carrera ingresada para el usuario: "+ new_student['USERNAME'][i] +" no existe"})
                
                for i in range(len(new_student)):
                    if len(str(new_student['TELEFONO'][i])) == 7 or len(str(new_student['TELEFONO'][i])) == 10:
                        conteo = 0 + conteo
                        
                    else:
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El usuario "+ new_student['USERNAME'][i] +" debe tener 7(fijo) o 10(celular) números, usted ingresó: "+ str(new_student['TELEFONO'][i])})
                
                for i in range(len(new_student)):
                    ced= str(new_student['CEDULA'][i])
                    if ced[0] == 0 :
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El usuario "+ new_student['USERNAME'][i] +" tiene un numero de documento que inicia por cero, por favor valide."})
                    
                    else:
                        conteo = 0 + conteo
                
                for i in range(len(new_student)):
                    ced= str(new_student['CEDULA_ACUDIENTE'][i])
                    if ced[0] == 0 :
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El acudiente del usuario "+ new_student['USERNAME'][i] +" tiene un numero de documento que inicia por cero, por favor valide."})
                    
                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    if len(str(new_student['TELEFONO_ACUDIENTE'][i])) == 7 or len(str(new_student['TELEFONO_ACUDIENTE'][i])) == 10:
                        conteo = 0 + conteo
                    else:
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El teléfono del acudiente del usuario "+ new_student['USERNAME'][i] +" debe tener 7(fijo) o 10(celular) números, usted ingresó: "+ str(new_student['TELEFONO_ACUDIENTE'][i])})
                
                for i in range(len(new_student)):
                    if (fecha_actual.year - new_student['NACIMIENTO'][i].year) <= 13:
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El usuario "+ new_student['USERNAME'][i] +" no se puede matricular, debe de tener más de 13 años"})

                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    if not CatalogsTypesDocuement.objects.filter(nombre=new_student['T_DOCUMENTO'][i]):
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El usuario "+ new_student['USERNAME'][i] +" tiene un tipo de documento incorrecto: " + new_student['T_DOCUMENTO'][i] + ", seleccione uno de la lista desplegable."})
                    else:
                        conteo = 0 + conteo
                
                for i in range(len(new_student)):
                    if not CatalogsSede.objects.filter(sede=new_student['SEDE'][i]):
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El usuario "+ new_student['USERNAME'][i] +" tipo de sede incorrecta" + new_student['SEDE'][i]+ ", seleccione uno de la lista desplegable."})

                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    if not re.search(r"^[A-Za-z0-9_!#$%&'*+\/=?`{|}~^.-]+@[A-Za-z0-9.-]+$", new_student['EMAIL'][i]):
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El usuario "+ new_student['USERNAME'][i] + " tiene datos que no son emails " + new_student['EMAIL'][i]})

                    else:
                        conteo = 0 + conteo
                
                if len(mensaje1) > 0:
                    response = JsonResponse(mensaje1, safe= False)
                    response.status_code = 400
                    return response

                # si pasa las validaciones, se comienza a procesar
                else:
                    for i in range(len(new_student)):
                        carrera = Programas.objects.get(programa_name=new_student['CARRERA_ID'][i]).id
                        costo_cierre = Programas.objects.get(programa_name=new_student['CARRERA_ID'][i]).costo
                        document = CatalogsTypesDocuement.objects.get(nombre=new_student['T_DOCUMENTO'][i]).id
                        sede = CatalogsSede.objects.get(sede=new_student['SEDE'][i]).id
                        data_list.append(Estudiante( codigo = int(codigo) + i, cedula= new_student['CEDULA'][i],
                                    nombre =new_student['NOMBRE'][i], apellidos= new_student['APELLIDOS'][i], 
                                    nacionalidad = new_student['NACIONALIDAD'][i],
                                    telefono =new_student['TELEFONO'][i], direccion= new_student['DIRECCION'][i],
                                    nacimiento = new_student['NACIMIENTO'][i],email=new_student['EMAIL'][i], 
                                    sexo = new_student['SEXO'][i], username= new_student['USERNAME'][i],
                                    nombre_acudiente =new_student['NOMBRE_ACUDIENTE'][i], apellidos_acudiente= new_student['APELLIDOS_ACUDIENTE'][i],
                                    telefono_acudiente = new_student['TELEFONO_ACUDIENTE'][i],cedula_acudiente=new_student['CEDULA_ACUDIENTE'][i], 
                                    carrera_id = carrera, masivo= True, pensum_asig_id = 1, periodo_matriculado_id= 1, tDocument_id = document,
                                    sede_id=sede, costo_cierre=costo_cierre,
                        ))
                        User.objects.create_user(codigo = int(codigo) + i, nombres =new_student['NOMBRE'][i], apellidos= new_student['APELLIDOS'][i],
                            email=new_student['EMAIL'][i],username= new_student['USERNAME'][i], tipe = CatalogsTypesRol.objects.get(rol="Estudiante"), 
                            password= Estudiante.objects.get_secret("RANDOM"), is_superuser=False, is_active = True, is_staff=False
                            
                            )

                    Estudiante.objects.bulk_create(data_list)

                    
                    #Generamos las facturas relacionadas con el estudiante, hace referencia a matrícula y a las facturas de las pensiones mensuales
                    
                    
                    for i in range(len(new_student)):
                        codigo_invoice= Facturas.objects.code_invoice()
                        for e in l:
                            data_fac_matricula.append(Facturas(
                                    user=User.objects.get(username=new_student['USERNAME'][i]),
                                    codigo=int(codigo_invoice) + i,
                                    email=new_student['EMAIL'][i],
                                    monto =Programas.objects.get(programa_name=new_student['CARRERA_ID'][i]).matricula,
                                    descripcion= e,
                                    estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),

                            ))
                        for e in range(int(Programas.objects.get(programa_name=new_student['CARRERA_ID'][i]).cuotas)):
                            data_fac_matricula.append(Facturas(
                                user=User.objects.get(username=new_student['USERNAME'][i]),
                                codigo=int(codigo_invoice) + i + e + 1,
                                email=new_student['EMAIL'][i],
                                monto =Programas.objects.get(programa_name=new_student['CARRERA_ID'][i]).cuota_valor,
                                descripcion="Mensualidad número "+ str(e+1),
                                estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),

                            ))
                        Facturas.objects.bulk_create(data_fac_matricula)

                    return HttpResponseRedirect(
                        reverse(
                        'academico_app:list-student'
                        )
                    )
        else:
            mensaje = 'Archivo inválido, recuerde que el archivo tiene por nombre "cargue_estudiantes.xlsx" , por favor verifique y cárguelo nuevamente'
            response = JsonResponse({"error": mensaje})
            response.status_code = 400
            return response
    
                
#Vista para generar plantilla que luego sirve para el cargue de las notas. Se encuentra ok
class Export_plant_est_csv(View):


    def post(self, request, *args, **kwargs):

        corte = self.request.POST.get('corte')
        materia = Banner.objects.filter(
            materia_id= self.request.POST.get('materia')
        )
        
        wb= Workbook()
        ws1 = wb.create_sheet(index=0, title="Plantilla")
        for number in range(0, len(CORTE)):
            c1 = ws1.cell(row=1, column= number + 1 )
            c1.value = CORTE[number]+ "_" + str(corte)
                
                   
        var_est = 2
        for number in materia:
            c1 = ws1.cell(row= var_est, column= 1 )
            c1.value = number.cod_student.codigo
            c1 = ws1.cell(row= var_est, column= 2 )
            c1.value = str(number.cod_student)
            c1 = ws1.cell(row= var_est, column= 3 )
            c1.value = str(number.materia.codigo)
            var_est += 1

        content = save_virtual_workbook(wb)
        response = HttpResponse(content)
        response['Content-Disposition'] = 'attachment; filename=cargue_notas_estudiantes.xlsx'
        response['Content-Type'] = 'application/x-xlsx'
        return response

#Vista para generar plantilla de estudiantes, se encuentra ok
def export_users_csv(request):

    #Variables utilizadas para traerse la información creada en modelos y catálogos
    program = Programas.objects.all()
    documents = CatalogsTypesDocuement.objects.all()
    sede = CatalogsSede.objects.all()
    wb = Workbook()

    #Nombramos las pestanas que va a llevar el archivo
    ws1 = wb.create_sheet(index=0, title="Plantilla")
    ws = wb.create_sheet(index=1, title="Campos")

    #Generamos los datos que van a las celdas, y que sirven de listas desplegables
    for number in range(0,len(program)): 
        ws['A{}'.format(number+1)].value= "{}".format(program[number])
    
    for number in range(0,len(COUNTRIES)): 
        ws['B{}'.format(number+1)].value= "{}".format(COUNTRIES[number][1])

    for number in range(0,len(GENEROS)): 
        ws['C{}'.format(number+1)].value= "{}".format(GENEROS[number][1])

    for number in range(0,len(documents)): 
        ws['D{}'.format(number+1)].value= "{}".format(documents[number])

    for number in range(0,len(sede)): 
        ws['E{}'.format(number+1)].value= "{}".format(sede[number])

    #Asignamos los nombres de los encabezados a las columnas del archivo de excel
    for number in range(0, len(EST_STUDENT)):
        c1 = ws1.cell(row=1, column= number + 1)
        c1.value = EST_STUDENT[number]
    
    #Asignamos las listas de validación para el excel
    data_val1 = DataValidation(type="list",formula1='=Campos!$A$1:$A$' + str(len(program))) 
    data_val2 = DataValidation(type="list",formula1='=Campos!$B$1:$B$' + str(len(COUNTRIES)))
    data_val3 = DataValidation(type="list",formula1='=Campos!$C$1:$C$' + str(len(GENEROS)))
    data_val4 = DataValidation(type="list",formula1='=Campos!$D$1:$D$' + str(len(documents)))
    data_val5 = DataValidation(type="list",formula1='=Campos!$E$1:$E$' + str(len(sede)))

    #Asignamos las listas desplegables al la hoja principal
    ws1.add_data_validation(data_val1)
    ws1.add_data_validation(data_val2)
    ws1.add_data_validation(data_val3)
    ws1.add_data_validation(data_val4)
    ws1.add_data_validation(data_val5)

    data_val1.add(ws1["P2"])
    data_val2.add(ws1["E2"])
    data_val3.add(ws1["J2"])
    data_val4.add(ws1["A2"])
    data_val5.add(ws1["Q2"])
    

    content = save_virtual_workbook(wb)
    response = HttpResponse(content)
    response['Content-Disposition'] = 'attachment; filename=cargue_estudiantes.xlsx'
    response['Content-Type'] = 'application/x-xlsx'
    return response


#Con esta vista mostramos los estudiantes aptos para graduarse, se encuentra ok
class ListFinalily(ListView):
    model=Banner
    second_model = Programas
    template_name = 'Academico/Estudiantes/list_aptos.html'
    context_object_name = 'finaly'

    def get_queryset(self):
        estudiante_list = []
        aceptado = Banner.objects.filter( promedio__gt= 2.99).values('cod_student').annotate(total=Count('cod_student')).order_by('total')
        
        for i in list(aceptado):
            st=Estudiante.objects.get(id=i['cod_student'])
            if Graduated.objects.filter(student_id = st.pk).count() == 0:
               carrera = Programas.objects.get(id=st.carrera_id)
               pagadas = Facturas.objects.filter(user_id=st.codigo, estado_id=2).count()
               rr={'pk':st.pk, 'codigo': st.codigo,'estudiante':st.apellidos +' '+ st.nombre , 'programa': carrera.programa_name , 'indicador': int((int(i['total'])/int(carrera.aceptado))*100),'financiero': int((int(pagadas)/(2 + int(carrera.cuotas)))*100) }
            
               estudiante_list.append(rr)
        def get_age(indica):
            return indica.get('indicador')
        
        estudiante_list.sort(key=get_age, reverse=True)
        return estudiante_list

class GraduateView(CreateView):
    model = Graduated
    form_class = GraduateRegisterForm
    template_name = 'Academico/Estudiantes/list_aptos.html'
    success_url =  reverse_lazy('academico_app:list-top-graduado')

    def get(self, request, *args, **kwargs):
        form =GraduateRegisterForm()
        context = {"form" : form} 
        return render(request, r"Academico\Estudiantes\studentGraduated.html", context)
    
    def post(self, request, *args, **kwargs):
        todos = []
        date = datetime.now()
        studen_graduate = str(self.request.POST.get('concat'))
        studen_folio = str(self.request.POST.get('folio'))
        studen_libro = str(self.request.POST.get('libro'))
        studen_graduate = studen_graduate.split(sep=",")

        for i in studen_graduate:
            consulta = Estudiante.objects.get(pk=i)
            individual= Graduated(student_id =int(i), carrera_id= int(consulta.carrera_id), periodo_grado=str(date.year) + "-" + Graduated.objects.mes(date), libro=studen_libro, folio=studen_folio)
            todos.append(individual)
            Estudiante.objects.filter(id=int(i)).update(is_active = False)
        Graduated.objects.bulk_create(todos)
        messages.success(self.request, 'Perfecto, graduaste estudiantes de manera exitosa')
        return HttpResponseRedirect(
                self.request.META.get("HTTP_REFERER")
                
        )

class StudentDeleteView(DeleteView):
    template_name = 'Academico/Estudiantes/delete_student.html'
    model = Estudiante
    second_model = User
    success_url = reverse_lazy('academico_app:list-student')

class TeacherDeleteFinalView(View):

    def post(self, request, pk, *args, **kwargs):
        consulta1= Estudiante.objects.get(codigo=pk)
        consulta= Banner.objects.filter(cod_student_id = consulta1.id).count()
        
        if consulta == 0:
            object1=Estudiante.objects.filter(codigo=pk).delete()
            object2=User.objects.filter(pk=pk).delete()
        
        else:
            j = 0 
            consulta_est = Banner.objects.filter(cod_student_id = consulta1.id)
            for i in consulta_est:
                suma = i.corte1 + i.corte2 + i.corte3  
                j = j + suma
            print(j)
            if j > 0:
               object1=Estudiante.objects.filter(codigo=pk).update(is_active = False)
               object2=User.objects.filter(pk=pk).update(is_active = False)
            else:
                object1=Estudiante.objects.filter(codigo=pk).delete()
                object2=User.objects.filter(pk=pk).delete()
                object3=Banner.objects.filter(cod_student_id = consulta1.id).delete()
            
            
        return HttpResponseRedirect( reverse_lazy('academico_app:list-student'))
    
class AsignamateriasCreateView(View):
    
    def post(self, request, *args, **kwargs):

        estudiante = str(self.request.POST.get('estudiante'))
        carrera = str(self.request.POST.get('carrera'))
        materias = str(self.request.POST.get('concat')) 
        materias=materias.split(sep=",")
        if materias[0] != "":
           for j in materias:
               Banner.objects.create(materia_id=int(j), cod_student_id= int(estudiante),
               Programa_id= int(carrera), corte1=0.0, corte2=0.0, corte3=0.0, promedio=0.0)

           Estudiante.objects.filter(id=int(estudiante)).update(is_matriculado=True) 
           return HttpResponseRedirect(
                    reverse(
                    'academico_app:list-student'
                    )
                )
        else:
            messages.warning(self.request, 'Por Favor seleccione por lo menos una materia para asignar al estudiante')
            return HttpResponseRedirect(
                    self.request.META.get("HTTP_REFERER")
                    
            )

class AddMateriasList(ListView):
    model= Banner
    template_name = 'Academico/Estudiantes/AddMateriasList.html'
    paginate_by = 10
    context_object_name = 'agregando'

    def get_context_data(self, **kwargs):
        cod = self.kwargs['pk']
        context = super(AddMateriasList, self).get_context_data(**kwargs)
        context['student'] = Estudiante.objects.filter(pk=cod)
        context['materias'] = Banner.objects.filter(cod_student=cod)
        return context
    

    def get_queryset(self):
        base = self.kwargs['pk']
        data_matriculadas = []
        data_materias = []
        data_estudiante = []

        ### periodos del alumno
        for i in  Estudiante.objects.filter(id=base):
            e = [i.id, i.carrera_id, i.periodo_matriculado_id]
            data_estudiante.append(e)
        
        #### materias banner
        for i in Banner.objects.filter(cod_student_id=base):
            data_matriculadas.append(i.materia_id)

        for i in Materias.objects.filter(periodo_id = int(data_estudiante[0][2]), programa_id= int(data_estudiante[0][1])):
            
            data_materias.append(i.id)
        
        for i in data_matriculadas:
            data_materias.remove(i)
        try:
            materias_final =  Materias.objects.filter(periodo_id = data_estudiante[0][2], programa_id= data_estudiante[0][1], id__in= [x for x in data_materias])
        except:
            materias_final = "No hay más materias por asignar"
        return  materias_final



class GraduatedListView(ListView):
    model= Graduated
    template_name = 'Academico/Estudiantes/GraduadoList.html'
    paginate_by = 15
    context_object_name = 'graduados'

class HabiliteTemplateView(ListView):
    template_name = "Academico/settings/Habilitaciones/permitir.html"
    model= Habilitaciones
    context_object_name = 'data'

    def get_context_data(self, **kwargs):
        context = super(HabiliteTemplateView, self).get_context_data(**kwargs)
        data = Habilitaciones.objects.latest("id").id
        context['habilitados'] = Habilitaciones.objects.filter(id=data)

        return context

    def get_queryset(self):
        
        consulta = Acciones.objects.all().order_by("-id")
        return consulta
    
# Vista que genera y habilita los cortes, se encuentra ok
class GenerateCorteView(View):
    
    def post(self, request, *args, **kwargs):

        data = str(self.request.POST.get('data'))
        data_cierre = Habilitaciones.objects.latest("id")
        informacion = User.objects.get(username=self.request.user)
        

        if data == "generar":

            if data == "generar" and data_cierre.Cerrado == True :

                Habilitaciones.objects.create(C1=True, C2=False, C3=False, Cerrado = False, B1=False, B2=False, B3=False, CF=False)
                Acciones.objects.create(usuario = str(informacion.username), nombre= str(informacion.nombres) +" "+ str(informacion.apellidos), accion = "Ha generado nuevos cortes")
              
                return HttpResponseRedirect(
                        reverse(
                        'academico_app:habilitado'
                        )
                    )
        
            else:
                messages.warning(self.request, 'No puede generar nuevos periodos teniendo aun periodos activos')
                return HttpResponseRedirect(
                        reverse(
                        'academico_app:habilitado'
                        )
                    )

        elif data == "C1":
            if data_cierre.C1 == True and data == "C1" and data_cierre.B1 == False :
                Habilitaciones.objects.filter(id=data_cierre.id).update(C1=False, B1= True )
                Acciones.objects.create(usuario = str(informacion.username), nombre= str(informacion.nombres) +" "+ str(informacion.apellidos), accion = "Ha habilitado el corte 1")
                return HttpResponseRedirect(
                        reverse(
                        'academico_app:habilitado'
                        )
                    )
            else:
                Habilitaciones.objects.filter(id=data_cierre.id).update(C1=True, C2=True, B1= True )
                Acciones.objects.create(usuario = str(informacion.username), nombre= str(informacion.nombres) +" "+ str(informacion.apellidos), accion = "Ha cerrado el corte 1")
                return HttpResponseRedirect(
                        reverse(
                        'academico_app:habilitado'
                        )
                    )


        elif data == "C2":
            if data_cierre.C2 == True and data == "C2" and data_cierre.B2 == False :
                Habilitaciones.objects.filter(id=data_cierre.id).update(C2=False, B2= True )
                Acciones.objects.create(usuario = str(informacion.username), nombre= str(informacion.nombres) +" "+ str(informacion.apellidos), accion = "Ha habilitado el corte 2")
                
                return HttpResponseRedirect(
                        reverse(
                        'academico_app:habilitado'
                        )
                    )
            else:
                Habilitaciones.objects.filter(id=data_cierre.id).update(C2=True, C3=True, B2= True )
                Acciones.objects.create(usuario = str(informacion.username), nombre= str(informacion.nombres) +" "+ str(informacion.apellidos), accion = "Ha cerrado el corte 2")
                return HttpResponseRedirect(
                    reverse(
                    'academico_app:habilitado'
                    )
                )
        
        elif data == "C3":

            if data_cierre.C3 == True and data == "C3" and data_cierre.B3 == False :
                Habilitaciones.objects.filter(id=data_cierre.id).update(C3=False, B3= True )
                Acciones.objects.create(usuario = str(informacion.username), nombre= str(informacion.nombres) +" "+ str(informacion.apellidos), accion = "Ha habilitado el corte 3")

                return HttpResponseRedirect(
                    reverse(
                    'academico_app:habilitado'
                    )
                )
            else:
                Habilitaciones.objects.filter(id=data_cierre.id).update(C3=True, CF=True, B3= True )
                Acciones.objects.create(usuario = str(informacion.username), nombre= str(informacion.nombres) +" "+ str(informacion.apellidos), accion = "Ha cerrado el corte 3")
                return HttpResponseRedirect(
                    reverse(
                    'academico_app:habilitado'
                    )
                )
        
        else:
            if data == "CF":
                Habilitaciones.objects.filter(id=data_cierre.id).update(CF=True, Cerrado = True)
                Banner.objects.filter(materia__is_active=True).update(promedio=(F('corte1')+F('corte2')+F('corte3'))/3)
                Materias.objects.filter(is_active=True).update(is_active=False, updated_at=datetime.now() )
                Acciones.objects.create(usuario = str(informacion.username), nombre= str(informacion.nombres) +" "+ str(informacion.apellidos), accion = "Ha cerrado el semestre")
                return HttpResponseRedirect(
                    reverse(
                    'academico_app:habilitado'
                    )
                )


#Vista ok aunque se puede seguir mejorando, cargue de notas de los estudiantes a la BD
class UpdateNotasView(View):

    def get(self, request, *args, **kwargs):
        form = UploadForm()
        context = {"form" : form} 
        return render(request, r"Academico\Docentes\upload_template.html", context)
    
    def post(self, request, *args, **kwargs):

        #Obtenemos datos del archivo cargado por el usuario y adicionalmente obtenemos el valor del input hidden
        nota_student = self.request.FILES["archivo"]
        corte = "_"+str(self.request.POST.get('corteFull'))
        asignatura = self.request.POST.get('asignatura')
         
        
        #Asignación de variables y creaciones de listas para verificar la información que cargó el usuario.
        conteo = 0
        conteo2 = 0
        conteo3 = 0
        conteo4 = 0
        student_no_create =[]
        notas_no_create =[]
        notas_no_decimal =[]
        notas_no_asignatura =[]

        #Validación 1: Verificamos que el documento cargado si se llame como es debido
        if nota_student.name != 'cargue_notas_estudiantes.xlsx' :
            mensaje = 'El archivo que intenta cargar no es el correcto.'
            response = JsonResponse({"error": mensaje})
            response.status_code = 400
            return response
        
        #Validación 2: Verificamos datos del archivo
        else:

            nota_student = pd.read_excel(nota_student, sheet_name="Plantilla", engine='openpyxl' )
            miss_values_count = nota_student.isnull().sum()
            miss_values_count = miss_values_count[miss_values_count != 0]

            for i, a in zip(list(nota_student), CORTE):
                if i != a + corte:
                    break

            # zona de verificación de datos en el archivo, datos en blanco
            if miss_values_count.shape[0] or i != a + corte :
                if  miss_values_count.shape[0]:
                    for name, miss_vals in miss_values_count.items():
                        p = miss_vals > 1
                        mensaje = f"  - A la columna '{name}' le falta{'n' if p else ''} " f"{miss_vals} dato{'s' if p else ''}."
                        response = JsonResponse({"error": mensaje})
                        response.status_code = 400
                        return response

                else:
                    mensaje = 'La columna con encabezado '+ i + ' en su archivo, no es válido, verifique el archivo y vuelva a cargarlo'
                    response = JsonResponse({"error": mensaje})
                    response.status_code = 400
                    return response

            else:
                #Verificamos que:  1. exista el estudiante en el salon, 2. las notas se encuentren entre cero y cinco.
                # 3. No tener calificaciones en enteros, 4 que el código de la asignatura si sea el correcto
                for i in range(len(nota_student)):
                    try: 
                        resultado = Banner.objects.get(cod_student_id = Estudiante.objects.get(codigo = str(nota_student['CÓDIGO'+ corte][i])).id, 
                        materia_id = Materias.objects.get(codigo = str(nota_student['COD_MATERIA'+ corte][i])).id)
                        conteo = 0 + conteo
                    except:
                        conteo = 1 + conteo
                        student_no_create.append(nota_student['NOMBRE'+ corte][i])
                
                for i in range(len(nota_student)):
                    try: 
                        float(nota_student['CORTE'+ corte][i])
                        conteo3 = 0 + conteo3
                        if float(nota_student['CORTE'+ corte][i]) < 0.0 or float(nota_student['CORTE'+ corte][i]) > 5.0:
                            conteo2 = 1 + conteo2
                            notas_no_create.append(nota_student['NOMBRE'+ corte][i]) 
                    
                        else:
                            conteo2 = 0 + conteo2
                    
                    except:
                        conteo3 = 1 + conteo3
                        notas_no_decimal.append(nota_student['NOMBRE'+ corte][i])
                
                for i in range(len(nota_student)):
                    if str(nota_student['COD_MATERIA'+ corte][i]) != str(asignatura):
                        conteo4 = 1 + conteo4
                        notas_no_asignatura.append(asignatura) 
                    else:
                        conteo4 = 0 + conteo4

                if conteo > 0:
                    for i in student_no_create:
                        mensaje = "Los datos introducidos en el archivo son erróneos, intenta cargar notas a alumnos que no existen en esta asignatura. Vuelva a descargar el archivo y rellene nuevamente los datos. Nota: el archivo que se genera corresponde exclusivamente a la asignatura en pantalla "
                        response = JsonResponse({"error": mensaje})
                        response.status_code = 400
                        return response
                
                elif conteo4 > 0:
                    for i in notas_no_asignatura:
                        mensaje = "Está intentando cargar notas a una asignatura que no corresponde, recuerde que en este caso se deben cargar notas a la asignatura con código: " + str(i)
                        response = JsonResponse({"error": mensaje})
                        response.status_code = 400
                        return response
                    
                elif conteo3 > 0:
                    for i in notas_no_decimal:
                        mensaje ="El Estudiante " + str(i) + " contiene valores en notas que no son decimales, por favor verifique."
                        response = JsonResponse({"error": mensaje})
                        response.status_code = 400
                        return response

                elif conteo2 > 0:
                    for i in notas_no_create:
                        mensaje ="El Estudiante " + str(i) + " contiene notas que no se encuentran bajo los parámetros, recuerde que deben estar en un rango de 0.0 a 5.0"
                        response = JsonResponse({"error": mensaje})
                        response.status_code = 400
                        return response
                
                    
                #4. Actualizamos las notas del estudiante.    
                else:
                    for i in range(0, len(nota_student)):
                        
                        if corte == "_1":
                           actualizar = Banner.objects.filter(cod_student_id = int(Estudiante.objects.get(codigo = nota_student['CÓDIGO'+ corte][i]).id),
                            materia_id = int(Materias.objects.get(codigo = nota_student['COD_MATERIA'+ corte][i]).id)).update(corte1= float(nota_student['CORTE_1'][i]))

                        elif corte == "_2":
                           actualizar = Banner.objects.filter(cod_student_id = int(Estudiante.objects.get(codigo = nota_student['CÓDIGO'+ corte][i]).id),
                            materia_id = int(Materias.objects.get(codigo = nota_student['COD_MATERIA'+ corte][i]).id)).update(corte2=float(nota_student['CORTE_2'][i]))
                    
                        else:
                           actualizar = Banner.objects.filter(cod_student_id = int(Estudiante.objects.get(codigo = nota_student['CÓDIGO'+ corte][i]).id),
                            materia_id = int(Materias.objects.get(codigo = nota_student['COD_MATERIA'+ corte][i]).id)).update(corte3= float(nota_student['CORTE_3'+ corte][i]))

                    return HttpResponseRedirect(
                        self.request.META.get("HTTP_REFERER") 
                    )

                    




        


        


    

