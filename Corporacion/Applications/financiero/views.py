from django.shortcuts import render, HttpResponseRedirect
from django.http import JsonResponse, HttpResponse
from django.views.generic import (TemplateView,
                                  FormView, CreateView, DeleteView, UpdateView,
                                  DetailView, ListView, View
                                  )
from django.db.models import Count, Sum
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models.functions import Coalesce
from django.db.models import FloatField


from Applications.academico.models import *
from .models import *
from .forms import *

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
import pandas as pd


# Vista que muestra el contexto de la factura a mostrar, se encuentra ok
class InvoiceListGeneral(ListView):
    model = Facturas
    template_name = 'Financiero/Listinvoicesgeneral.html'
    context_object_name = 'invoice'

    def get_context_data(self, **kwargs):
        context = super(InvoiceListGeneral, self).get_context_data(**kwargs)
        usuario = list(Facturas.objects.all().values("user").distinct())
        datos = []

        for i in usuario:

            pendiente = Facturas.objects.filter(
                user=i['user'], estado_id=1).count()
            pagadas = Facturas.objects.filter(
                user=i['user'], estado_id=2).count()
            abono = Facturas.objects.filter(
                user=i['user'], estado_id=3).count()
            data_final = {"Codigo": str(User.objects.get(codigo=i['user']).codigo), "Usuario": str(User.objects.get(
                codigo=i['user']).nombres) + " "+str(User.objects.get(codigo=i['user']).apellidos), "Pagadas": pagadas, "Pendientes": pendiente, "Abonadas": abono}
            datos.append(data_final)

        context['student'] = datos
        return context


# Vista que muestra los pagos de las mensualidades y pagos obligatorios del estudiante, se encuentra ok
class InvoiceListviewUser(ListView):
    model = Facturas
    template_name = 'Financiero/Listinvoices.html'
    context_object_name = 'invoice'

    def get_context_data(self, **kwargs):
        context = super(InvoiceListviewUser, self).get_context_data(**kwargs)
        context["datos"] = Estudiante.objects.filter(pk=self.kwargs['pk'])
        context['progreso'] = int(int(Facturas.objects.filter(user_id=Estudiante.objects.get(pk=self.kwargs['pk']).codigo, estado_id=2).count())/int(Programas.objects.get(
            id=Estudiante.objects.get(pk=self.kwargs['pk']).carrera_id
        ).cuotas+2)*100)
        return context

    def get_queryset(self):
        info = []
        datos = Facturas.objects.filter(user_id=Estudiante.objects.get(
            pk=self.kwargs['pk']).codigo).order_by("codigo")
        for i in datos:
            pagado = FacturasSub.objects.filter(
                facturas_id=i.pk).aggregate(pagados=Sum("pagado"))
            pagados_data = FacturasSub.objects.filter(facturas_id=i.pk)
            data = pagado['pagados']
            datos_final = {'pk': i.pk, "codigo": i.codigo, 'descripcion': i.descripcion,
                           'estado': i.estado, 'monto': f'$ {i.monto:,.2f}', 'pagado':  f'$ {data if data else 0.0 :,.2f}'}

            info.append(datos_final)
        return info


class ListInvoiceDetailView(View):

    def get(self, request, *args, **kwargs):
        info = []
        pagados_data = FacturasSub.objects.filter(
            facturas_id=Facturas.objects.get(id=int(self.request.GET.get('info'))))

        for e in pagados_data:
            datos = {"pk": e.id, 'consecutivo': e.consecutivo, 'observacion': e.observacion, 'payed': f'$ {e.pagado:,.2f}', 'fecha': str(
                e.created_at.day) + "-" + str(e.created_at.month) + "-" + str(e.created_at.year)}
            info.append(datos)
        response = {}
        response['data'] = info
        return JsonResponse(response)


class InvoiceDetailView(FormView):

    template_name = 'Financiero/update_invoice.html'
    model = Facturas
    form_class = FacturasForm

    def get_context_data(self, **kwargs):
        context = super(InvoiceDetailView, self).get_context_data(**kwargs)
        pagado = FacturasSub.objects.filter(
            facturas_id=self.kwargs['pk']).aggregate(pagados=Sum("pagado"))
        data = Facturas.objects.get(pk=self.kwargs['pk'])
        monto = data.monto
        pagado2 = pagado['pagados']
        pendiente = Facturas.objects.manejo(data, pagado)
        datos = []
        info = {"codigo": data.codigo, "monto": f'$ {monto:,.2f}', 'pagado': f'$ {pagado2 if pagado2 else 0.0 :,.2f}',
                'pendiente': f'$ {pendiente:,.2f}', 'pendiente2': pendiente}
        datos.append(info)
        context["datos"] = datos
        return context


# Vista que crea sub facturas cuando se hacen abonos, pagos, etc se encuentra ok
class InvoiceSubCreate(View):

    def post(self, request, *args, **kwargs):

        mensaje1 = []
        pendiente = self.request.POST.get('pendiente')
        facturas = self.request.POST.get('facturas')
        observacion = self.request.POST.get('observacion')
        consecutivo = self.request.POST.get('consecutivo')
        pagado = self.request.POST.get('pagado')
        characters = ",0"
        pendiente = ''.join(x for x in pendiente if x not in characters)

        if FacturasSub.objects.filter(consecutivo=consecutivo) or Gastos.objects.filter(consecutivo=consecutivo):
            mensaje1.append(
                {"error": 'El consecutivo que intenta crear ya existe, por favor verifique el número del recibo de pago.'})
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response

        else:
            matricula = FacturasSub.objects.create(
                facturas=Facturas.objects.get(codigo=facturas),
                observacion=observacion,
                consecutivo=consecutivo,
                pagado=pagado,
            )
            fact_act = FacturasSub.objects.filter(facturas_id=Facturas.objects.get(
                codigo=facturas).id).aggregate(Sum('pagado'))
            monto = Facturas.objects.get(codigo=facturas).monto
            if int(fact_act['pagado__sum']) == int(monto):
                Facturas.objects.filter(codigo=facturas).update(
                    estado_id=CatalogsTypesInvoices.objects.get(estado="Pagada"))
            elif int(fact_act['pagado__sum']) > 0 and int(fact_act['pagado__sum']) < int(monto):
                Facturas.objects.filter(codigo=facturas).update(
                    estado_id=CatalogsTypesInvoices.objects.get(estado="Abono"))
            else:
                Facturas.objects.filter(codigo=facturas).update(
                    estado_id=CatalogsTypesInvoices.objects.get(estado="Pendiente"))

            mensaje1.append({"mensaje": 'No hay error'})
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 201
            return response


class FacturasSubDetailView(DetailView):
    template_name = 'Financiero/DetailSub.html'
    model = FacturasSub


class FacturasSubDeleteView(DeleteView):
    template_name = 'Financiero/DeleteFactSub.html'
    model = FacturasSub

    def post(self, request, *args, **kwargs):

        data_delete = self.kwargs['pk']

        restar = FacturasSub.objects.get(id=data_delete).pagado
        facturas = FacturasSub.objects.get(id=data_delete).facturas_id
        fact_act = FacturasSub.objects.filter(
            facturas_id=facturas).aggregate(Sum('pagado'))
        fact_act = float(fact_act['pagado__sum']) - float(restar)
        monto = Facturas.objects.get(id=facturas).monto
        teacher_delete = FacturasSub.objects.filter(id=data_delete).delete()

        if fact_act > 0 and fact_act < float(monto):
            Facturas.objects.filter(id=facturas).update(
                estado_id=CatalogsTypesInvoices.objects.get(estado="Abono"))
        else:
            Facturas.objects.filter(id=facturas).update(
                estado_id=CatalogsTypesInvoices.objects.get(estado="Pendiente"))
        messages.success(
            self.request, 'La factura ha sido eliminada correctamente')
        return HttpResponseRedirect(
            self.request.META.get("HTTP_REFERER")
        )


class InformeView(View):

    def post(self, request, *args, **kwargs):
        estudiantes = str(self.request.POST.get('concat2'))
        select = self.request.POST.get('informe12')
        estudiantes = estudiantes.split(sep=",")
        listaDataAcademic = []
        listaDataFinance = []

        for i in estudiantes:
            infoAcademic = Banner.objects.filter(cod_student_id=i)
            infoFinance = Facturas.objects.filter(
                user_id=Estudiante.objects.get(id=i).codigo)

            for j in infoAcademic:
                dataAcademic = {
                    "Codigo": j.cod_student.codigo, "Nombre": j.cod_student.nombre + " " +
                    j.cod_student.apellidos, "Tipo": "Académico",
                    "Materia": j.materia.materia.nombre_materia, "Promedio": float(j.promedio)
                }
                listaDataAcademic.append(dataAcademic)

            for l in infoFinance:
                dataFinance = {
                    "Codigo": l.user.codigo, "Nombre": l.user.nombres + " " +
                    l.user.apellidos, "Tipo": "Financiero",
                    "Factura": l.codigo, "Estado": l.estado.estado
                }
                listaDataFinance.append(dataFinance)

        dfAcademic = pd.DataFrame(listaDataAcademic)
        dfFinance = pd.DataFrame(listaDataFinance)

        # Creacion del archivo
        if select == "1":  # informe Academico
            wb = Workbook()
            ws = wb.create_sheet(index=0, title="Academico")

            headsList = list(dfAcademic.columns.values)

            for number in range(0, len(headsList)):
                Heads = ws.cell(row=1, column=number + 1)
                Heads.value = headsList[number]

            var_est = 2
            for number in range(0, len(dfAcademic)):

                c1 = ws.cell(row=var_est, column=1)
                c1.value = dfAcademic["Codigo"][number]
                c2 = ws.cell(row=var_est, column=2)
                c2.value = dfAcademic["Nombre"][number]
                c3 = ws.cell(row=var_est, column=3)
                c3.value = dfAcademic["Tipo"][number]
                c4 = ws.cell(row=var_est, column=4)
                c4.value = dfAcademic["Materia"][number]
                c5 = ws.cell(row=var_est, column=5)
                c5.value = dfAcademic["Promedio"][number]
                var_est += 1
            content = save_virtual_workbook(wb)
            response = HttpResponse(content)
            response['Content-Disposition'] = 'attachment; filename=informe_Academico.xlsx'
            response['Content-Type'] = 'application/x-xlsx'
            return response

        else:
            # informe financiero

            wb = Workbook()
            ws = wb.create_sheet(index=0, title="Financiero")

            headsList = list(dfFinance.columns.values)

            for number in range(0, len(headsList)):
                Heads = ws.cell(row=1, column=number + 1)
                Heads.value = headsList[number]

            var_est = 2
            for number in range(0, len(dfFinance)):

                c1 = ws.cell(row=var_est, column=1)
                c1.value = dfFinance["Codigo"][number]
                c2 = ws.cell(row=var_est, column=2)
                c2.value = dfFinance["Nombre"][number]
                c3 = ws.cell(row=var_est, column=3)
                c3.value = dfFinance["Tipo"][number]
                c4 = ws.cell(row=var_est, column=4)
                c4.value = dfFinance["Factura"][number]
                c5 = ws.cell(row=var_est, column=5)
                c5.value = dfFinance["Estado"][number]
                var_est += 1
            content = save_virtual_workbook(wb)
            response = HttpResponse(content)
            response['Content-Disposition'] = 'attachment; filename=informe_Financiero.xlsx'
            response['Content-Type'] = 'application/x-xlsx'
            return response
